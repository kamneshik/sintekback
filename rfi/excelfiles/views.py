import datetime
import pytz
import win32com.client as win32
import win32print
import pythoncom
import csv
import textwrap

from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_GET
from django.contrib.auth.decorators import login_required
from rficrud.models import RFI, Inspector
from django.shortcuts import render, redirect, get_object_or_404
from datetime import date
import openpyxl
from openpyxl.utils import rows_from_range
import os
from .forms import RFIForm, NumberRangeForm

local_path_to_templates = 'excelfiles/all_excel_files/excel_templates/'

local_tz = pytz.timezone('Europe/Moscow')
today_date = date.today()
#today_date = date(2023, 8, 10)
current_month = today_date.strftime("%m_%B")
current_year = today_date.strftime("%Y")
PATH_TO_SAVE = f'excelfiles/all_excel_files/{current_year}/{current_month}/'


save_today_date = today_date.strftime('%d-%m-%Y')

base_path = "C:/Users/Maksim/PycharmProjects/Django_Rfi/rfi/"


@login_required
def process_excel_file(request):
    """Func that get excel file with RFI numbers from JV and set it to DataBase"""
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(filename=excel_file, read_only=True)
        ws = wb.active

        last_row = ws.max_row
        # print(last_row)
        req_date = ws['C2'].value
        if req_date:
            daily_rfis = RFI.objects.filter(date_of_inspection__date=req_date)
            for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=last_row), start=2):
                try:
                    desc_of_work = ws[f"H{row_index}"].value
                    desc_of_work = desc_of_work.split('#')[0].strip()
                    print(type(desc_of_work), row_index, desc_of_work)
                    temporary_rfi = daily_rfis.get(description_of_work_russian=desc_of_work)
                    print(temporary_rfi.description_of_work_english)
                    number = ws[f'A{row_index}'].value[-5:]
                    temporary_rfi.rfi_number_from_akkuyu = int(number)
                    temporary_rfi.save()
                except:
                    pass
            messages.success(request, f"The operation was completed with Success!")
            return redirect('/success')

    return render(request, 'excelfiles/fill_rfi_numbers.html')


def success_page(request):
    return render(request, 'excelfiles/success.html')


def excel_file_to_browser(request):
    return render(request, 'excelfiles/excel_file_to_browser.html')


@require_GET
def get_rfis(request):
    """Функция возвращает JSON всех ЗНО

       СДЕЛАТЬ!!!
            фильтрацию перед выдачей, дать поля для выдачи
        """
    rfis = RFI.objects.all()
    data = []
    for rfi in rfis:
        data.append({
            'id': rfi.id,
            'excel_number': rfi.excel_number,
            # #'status': rfi.status,
            # 'rfi_number_from_akkuyu': rfi.rfi_number_from_akkuyu,
            # 'date_of_create': rfi.date_of_create.strftime('%Y-%m-%d'),
            # 'date_of_inspection': rfi.date_of_inspection.strftime('%Y-%m-%d'),
            # 'object_name': rfi.object_name.name,
            # 'project': rfi.project,
            # 'type_of_work': rfi.type_of_work,
            # 'description_of_work_russian': rfi.description_of_work_russian,
            # 'description_of_work_english': rfi.description_of_work_english,
            # 'quality_plan_code': rfi.quality_plan_code,
            # 'akkuyu_signer': rfi.akkuyu_signer.name,
            # 'independent_control': rfi.independent_control.name,
            # 'qc_signer': rfi.qc_signer.name,
            # 'design_supervision_signer': rfi.design_supervision_signer.name,
            # 'contractor': rfi.contractor.name if rfi.contractor else '',
            # 'contractor_sintek': rfi.contractor_sintek.name if rfi.contractor_sintek else '',
        })
    return JsonResponse(data, safe=False)


def from_django_to_excel(request):
    #today_date = date(2023, 6, 10)
    rfi_list = RFI.objects.filter(date_of_create__date=today_date)
    excel_template_path = local_path_to_templates + 'from_django_to_excel_template.xlsx'

    try:
        workbook = openpyxl.load_workbook(excel_template_path)
        worksheet = workbook.active

        row = 2
        row_height = 100
        for rfi in rfi_list:
            check_inspector(rfi)
            worksheet[f'C{row}'] = rfi.object_name.object_name
            worksheet[f'D{row}'] = rfi.project_name.name()
            worksheet[f'E{row}'] = rfi.description_of_work_russian
            worksheet[f'F{row}'] = rfi.description_of_work_english
            worksheet[f'I{row}'] = rfi.date_of_create.date()
            worksheet[f'J{row}'] = rfi.date_of_inspection.date()
            localized_inspection_time = rfi.date_of_inspection.astimezone(local_tz)
            worksheet[f'K{row}'] = localized_inspection_time.time()
            localized_create_time = rfi.date_of_create.astimezone(local_tz)
            worksheet[f'L{row}'] = localized_create_time.time()
            worksheet.row_dimensions[row].height = row_height

            if rfi.rfi_number_from_akkuyu:
                worksheet[f'H{row}'] = "ЗНО-" + str(rfi.rfi_number_from_akkuyu)
                # print(rfi.rfi_number_from_akkuyu, type(rfi.rfi_number_from_akkuyu))

            if rfi.is_active_beton:
                worksheet[f'G{row}'] = '+'

            row += 1

        print('HELLO GABE')

        folder_name = check_folder_exist()
        workbook.save(f'{folder_name}/From_django_to_Excel_{save_today_date}.xlsx')
        messages.success(request, f'File From_django_to_Excel_{save_today_date} Created')
    except FileNotFoundError:
        messages.error(request, f'File From_django_to_Excel_{save_today_date} failed')
        print('cant find file')

    # return redirect('print_page')
    return redirect('/rubka')


def dispathcher(request):
    form = NumberRangeForm()
    return render(request, 'excelfiles/dispathcher.html', {"form": form})


def print_page(request):
    """Начальная стадия контроллера для Создания и печати Excel файлов"""

    return render(request, 'excelfiles/print_page.html')


def create_json_answer(queryset):
    data = []
    for rfi in queryset:
        data.append({
            'id': rfi.id,
            'excel_number': rfi.excel_number,
            # #'status': rfi.status,
            # 'rfi_number_from_akkuyu': rfi.rfi_number_from_akkuyu,
            # 'date_of_create': rfi.date_of_create.strftime('%Y-%m-%d'),
            # 'date_of_inspection': rfi.date_of_inspection.strftime('%Y-%m-%d'),
            'object_name': rfi.object_name.object_name,
            # 'project': rfi.project.project_name,
            # 'type_of_work': rfi.type_of_work,
            # 'description_of_work_russian': rfi.description_of_work_russian,
            'description_of_work_english': rfi.description_of_work_english,
            # 'quality_plan_code': rfi.quality_plan_code,
            # 'akkuyu_signer': rfi.akkuyu_signer.name,
            # 'independent_control': rfi.independent_control.name,
            # 'qc_signer': rfi.qc_signer.name,
            # 'design_supervision_signer': rfi.design_supervision_signer.name,
            # 'contractor': rfi.contractor.name if rfi.contractor else '',
            # 'contractor_sintek': rfi.contractor_sintek.name if rfi.contractor_sintek else '',
        })
    return data


def create_request_list_tool(queryset, path_to_template):
    """Получает QuerySet, Путь до шаблона на основе этих данных формирует заявку на инспекцию
    """
    insp_date = queryset.last().date_of_inspection.strftime('%d-%m-%Y')
    try:
        print('Стартую')
        workbook = openpyxl.load_workbook(path_to_template)
        print('It works')
        worksheet = workbook.active
        # print(type(worksheet))

        row = 2
        row_height = 80
        #print(queryset)
        obj_name_dict = {}
        for rfi in queryset:
            print(rfi.project_name)
            print(f"Filing {row}'nd row")
            worksheet[f'B{row}'] = rfi.date_of_create.date()
            worksheet[f'C{row}'] = rfi.date_of_inspection.date()
            localozed_time = rfi.date_of_inspection.astimezone(local_tz)
            worksheet[f'D{row}'] = localozed_time.time()

            worksheet[f'E{row}'] = rfi.object_name.object_name

            worksheet[f'H{row}'] = rfi.description_of_work_russian + ' # ' + rfi.description_of_work_english
            worksheet[f'I{row}'] = rfi.project_name.quality_plan.quality_plan_name

            project = rfi.project_name
            worksheet[f'F{row}'] = f'{project.constriction_object}.' \
                                   f'{project.code}.{project.object_name}.' \
                                   f'{project.type_of_works}.{project.project_name}'
            worksheet[f'G{row}'] = rfi.object_name.object_place
            if rfi.pk_points:
                worksheet[f"J{row}"] = rfi.pk_points
            else:
                worksheet[f"J{row}"] = 'H-H-H-H'
            worksheet[f'M{row}'] = rfi.akkuyu_signer.signer_name
            worksheet[f'N{row}'] = rfi.independent_control.signer_name
            worksheet[f'O{row}'] = rfi.qc_signer.signer_name
            worksheet[f'P{row}'] = rfi.contractor.signer_name
            worksheet[f'R{row}'] = rfi.contractor_sintek
            worksheet[f'S{row}'] = rfi.design_supervision_signer.signer_name
            worksheet[f'V{row}'] = 'TSM ENERJI (Sintek)'
            worksheet.row_dimensions[row].height = row_height

            if rfi.object_name.object_name not in obj_name_dict:
                obj_name_dict[rfi.object_name.object_name] = 1

            copy_row_style(worksheet[2], worksheet[row])
            print(f'$$$$$$$$$$$$$$ row number {row}################')
            row += 1
        obj_str = ''
        for item in obj_name_dict:
            obj_str += item + '-'

        print('trying to save file')
        check_folder_exist(today_date)
        save_path = f'{PATH_TO_SAVE}/{save_today_date}/{obj_str}{insp_date}.xlsx'
        workbook.save(save_path)
        adjusting_all_rows(path=save_path)
        now = datetime.datetime.now()

        print(f'Request file saved {now}')
        message = f'Request for {obj_str}{insp_date} Created'
    except FileNotFoundError:
        message = f'Заявка на {insp_date} Failed'
        print('file problem')

    return message


def create_request_list(request):
    """Функция, которая создаёт заявку для получения номеров ЗНО"""

    rfi_list = RFI.objects.filter(date_of_create__date=today_date)
    rfi_list = rfi_list.filter(rfi_number_from_akkuyu=None)
    #print(rfi_list)

    rfi_list_dinara = rfi_list.filter(object_name__object_name__regex=r"00\w+")

    rfi_list_tatiana = rfi_list.filter(object_name__object_name__regex=r'^([1-9]\d|0[1-9])\w+$')
    #print(rfi_list_tatiana)
    excel_template_path = local_path_to_templates + 'request_template.xlsx'
    file_path = os.path.abspath(excel_template_path)
    try:
        print('1220')
        if rfi_list_dinara.exists():
            message = create_request_list_tool(rfi_list_dinara, file_path)
            messages.info(request, message)

        if rfi_list_tatiana.exists():
            message = create_request_list_tool(rfi_list_tatiana, file_path)
            messages.info(request, message)
    except:
        print('12')
    return redirect('/rubka')


def print_to_console_queryset(queryset):
    for item in queryset:
        print(item.object_name)


def points_from_pk(point_string):
    points = point_string.strip().split('-')
    return points[0], points[1], points[2], points[3]


def create_single_rfi(rfi_data):
    """Func that create only one rfi, Input data: rfi_data type:jQuery"""

    try:
        print('Trying to open file')
        workbook = openpyxl.load_workbook(local_path_to_templates + 'RFI_Template.xlsx')
        print('file opened')
        worksheet = workbook.active
        project = rfi_data.project_name
        worksheet[f'D4'] = f'{project.constriction_object}.' \
                           f'{project.code}.{project.object_name}.' \
                           f'{project.type_of_works}.{project.project_name}'
        if rfi_data.rfi_number_from_akkuyu:
            worksheet['D9'] = 'ЗНО-' + str(rfi_data.rfi_number_from_akkuyu)
        else:
            worksheet["D9"] = 'ЗНО-'
        worksheet['G9'] = rfi_data.date_of_create.date()
        localized_time = rfi_data.date_of_create.astimezone(local_tz)
        worksheet['G11'] = localized_time.time()
        worksheet['F13'] = rfi_data.date_of_inspection.date()
        localized_time = rfi_data.date_of_inspection.astimezone(local_tz)
        worksheet['F15'] = localized_time.time()
        worksheet['B13'] = rfi_data.object_name.object_place

        worksheet['A23'] = rfi_data.description_of_work_russian
        worksheet['A24'] = rfi_data.description_of_work_english
        worksheet['A28'] = project.quality_plan.quality_plan_name

        worksheet["C28"], worksheet["D28"], worksheet["E28"], worksheet["F28"] = points_from_pk(rfi_data.pk_points)
        worksheet['B38'] = rfi_data.akkuyu_signer.signer_name
        worksheet['C38'] = rfi_data.independent_control.signer_name
        worksheet['D38'] = rfi_data.design_supervision_signer.signer_name
        worksheet['E38'] = rfi_data.qc_signer.signer_name
        worksheet['F38'] = rfi_data.contractor.signer_name
        worksheet['G38'] = 'Серкан Кандемир /Serkan Kandemir\n +90 (552) 505-93-68'
        path_to_save = f"C:\\Users\Maksim\PycharmProjects\Django_Rfi\\rfi\excelfiles\excel_templates\\2023-04-25\RFI-{rfi_data.rfi_number_from_akkuyu}"

        if rfi_data.rfi_number_from_akkuyu:
            save_name = rfi_data.rfi_number_from_akkuyu
        else:
            save_name = 'Excel-Number' + rfi_data.excel_number
        workbook.save(f'{PATH_TO_SAVE}/{save_today_date}/print/RFI-{save_name}.xlsx')
    except FileNotFoundError:
        print('something wrong')


def create_single_checklist(rfi_data):
    """Func that create single checklist. Input data: rfi data type:jQuery
        Add variable check-lists
    """
    local_tz = pytz.timezone('Europe/Moscow')
    today_date = date.today()
    try:
        print('Trying to create CheckList file')
        workbook = openpyxl.load_workbook(local_path_to_templates + 'Check_list_template.xlsx')
        print('Checklist file opened')
        worksheet = workbook.active
        if rfi_data.rfi_number_from_akkuyu:
            worksheet['E4'] = 'ЗНО-' + str(rfi_data.rfi_number_from_akkuyu)
        else:
            worksheet["E4"] = 'ЗНО- '
        worksheet['E10'] = rfi_data.date_of_inspection.date()
        localozed_time = rfi_data.date_of_inspection.astimezone(local_tz)
        worksheet['E11'] = localozed_time.time()
        worksheet['E12'] = rfi_data.object_name.object_name

        project = rfi_data.project_name
        worksheet['E13'] = f'{project.constriction_object}.' \
                           f'{project.code}.{project.object_name}.' \
                           f'{project.type_of_works}.{project.project_name}'
        worksheet['E14'] = project.quality_plan.quality_plan_name
        worksheet['E15'] = rfi_data.object_name.object_place

        worksheet['E16'] = rfi_data.description_of_work_russian
        worksheet['E17'] = rfi_data.description_of_work_english

        worksheet['B45'] = 'Серкан Кандемир /Serkan Kandemir'
        worksheet['B48'] = rfi_data.qc_signer.signer_name
        if rfi_data.rfi_number_from_akkuyu:
            save_name = rfi_data.rfi_number_from_akkuyu
        else:
            save_name = 'Excel-Number' + rfi_data.excel_number
        workbook.save(f'{PATH_TO_SAVE}/{save_today_date}/print/CheckList-{save_name}.xlsx')
    except FileNotFoundError:
        print('something went wrong')


def create_single_checklist_new(rfi_data):
    """Func that create single checklist. Input data: rfi data type:jQuery
        Add variable check-lists
    """
    local_tz = pytz.timezone('Europe/Moscow')
    today_date = date.today()
    print("we in new checklist")
    try:
        print('Trying to create CheckList file')
        workbook = openpyxl.load_workbook(local_path_to_templates + 'Check_list_template_new.xlsx')
        print('Checklist file opened')
        worksheet = workbook.active
        if rfi_data.rfi_number_from_akkuyu:
            worksheet['D4'] = 'ЗНО-' + str(rfi_data.rfi_number_from_akkuyu)
        else:
            worksheet["D4"] = 'ЗНО- '
        worksheet['D8'] = rfi_data.date_of_inspection.date()
        localozed_time = rfi_data.date_of_inspection.astimezone(local_tz)
        worksheet['D9'] = localozed_time.time()
        worksheet['D10'] = rfi_data.object_name.object_name

        project = rfi_data.project_name
        worksheet['D11'] = f'{project.constriction_object}.' \
                           f'{project.code}.{project.object_name}.' \
                           f'{project.type_of_works}.{project.project_name}'

        worksheet['D12'] = rfi_data.object_name.object_place \
                           + "\n\n" + rfi_data.description_of_work_russian \
                           + "\n" + rfi_data.description_of_work_english

        # worksheet['E16'] = rfi_data.description_of_work_russian
        # worksheet['E17'] = rfi_data.description_of_work_english

        worksheet['C34'] = 'Серкан Кандемир /Serkan Kandemir'
        worksheet['C36'] = rfi_data.qc_signer.signer_name
        if rfi_data.rfi_number_from_akkuyu:
            save_name = rfi_data.rfi_number_from_akkuyu
        else:
            save_name = 'Excel-Number' + rfi_data.excel_number
        workbook.save(f'{PATH_TO_SAVE}/{save_today_date}/print/CheckList-{save_name}.xlsx')
    except FileNotFoundError:
        print('something went wrong')


def check_inspector(rfi):
    """"Принимает экземпляр модели RFI возвращает
        количство копий в зависимости от инспектора"""

    checklist_copies = 1
    rfi_copies = 2
    if rfi.inspector:
        checklist_copies = rfi.inspector.checklists_copies
        rfi_copies = rfi.inspector.rfi_copies

    if rfi.is_active_beton:
        checklist_copies = 0
        rfi_copies = 0

    return checklist_copies, rfi_copies


def copy_style(sample_cell, target_cell):
    target_cell.font = sample_cell.font.copy()
    target_cell.border = sample_cell.border.copy()
    target_cell.fill = sample_cell.fill.copy()
    target_cell.number_format = sample_cell.number_format
    target_cell.protection = sample_cell.protection.copy()
    target_cell.alignment = sample_cell.alignment.copy()


def copy_row_style(sample_row, target_row):
    for sample_cell, target_cell in zip(sample_row, target_row):
        copy_style(sample_cell, target_cell)


def adjusting_all_rows(path, cell_width=65, row_height = 14):
    """Функция которая подбирает оптимальную высоту строки на основании ячейки из столбца H.
     При заранее заданой ширине столбца и высоты одной строки """
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    index = 1
    for row in ws.iter_rows(min_row=2):
        cell = row[7]
        cell_text = str(cell.value)
        index += 1
        wrapped_text = textwrap.wrap(cell_text, width=cell_width)
        for text in wrapped_text:
            print(index, text, len(text))
        num_lines = len(wrapped_text)
        cell_height = row_height * num_lines
        ws.row_dimensions[cell.row].height = max(cell_height , 72)
    wb.save(path)


def create_daily_rfis(request, rfis_date=date.today()):
    """Func that create ALL daily RFIs, Input data: rfis_date type: Date"""
    rfi_list = RFI.objects.filter
    #today_date = date(2023, 6, 17)
    rfi_list = RFI.objects.filter(date_of_create__date=today_date)
    print(rfi_list)
    inscheduled_rfi_list = rfi_list.filter(date_of_inspection__date=today_date)
    scheduled_rfi_list = rfi_list.exclude(id__in=inscheduled_rfi_list.values('id'))

    check_folder_exist()
    tuples_list = []
    for rfi in scheduled_rfi_list:
        create_single_checklist_new(rfi)
        checklist_copies, rfi_copies = check_inspector(rfi)
        print(rfi.rfi_number_from_akkuyu)
        tuples_list.append((f"CheckList-{rfi.rfi_number_from_akkuyu}.xlsx", checklist_copies))

        create_single_rfi(rfi)
        tuples_list.append((f"RFI-{rfi.rfi_number_from_akkuyu}.xlsx", rfi_copies))

    for rfi in inscheduled_rfi_list:
        create_single_checklist_new(rfi)
        checklist_copies, rfi_copies = check_inspector(rfi)
        print(rfi.rfi_number_from_akkuyu)
        tuples_list.append((f"CheckList-Excel_number{rfi.excel_number}.xlsx", checklist_copies))

        create_single_rfi(rfi)
        tuples_list.append((f"RFI-Excel_number{rfi.excel_number}.xlsx", rfi_copies))

    tuples_to_csv(tuples_list, csv_file_path=f"{PATH_TO_SAVE}{save_today_date}/print/print_setting.csv")

    return redirect('/success')


def tuples_to_csv(tuples, csv_file_path):
    """Create print config file contain:Tuples of (file_name, copies_count)
        Input date tuples: List of tuples, csv_file_path"""

    with open(base_path + csv_file_path, 'w', newline='', encoding='utf-8') as csvfile:
        csv_writer = csv.writer(csvfile)
        for row in tuples:
            csv_writer.writerow(row)


def create_rfi(request):
    if request.method == 'POST':
        form = RFIForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('success_page')
    else:
        form = RFIForm()

    return render(request, 'excelfiles/create_rfi.html', {'form': form})


def check_folder_exist(check_date=date.today()):
    """Cheking exist folder for C"""
    final_date = check_date.strftime('%d-%m-%Y')
    folder_name = f'{PATH_TO_SAVE}/{final_date}/'
    print_folder = folder_name + "print"

    if not os.path.isdir(folder_name):
        os.makedirs(folder_name)

    elif not os.path.isdir(print_folder):
        os.makedirs(print_folder)

    return folder_name


def get_information(request):
    if request.method == "POST":
        form = NumberRangeForm(request.POST)
        if form.is_valid():
            print('get hish mazafaka')
            from_num = request.POST.get('from_number')
            to_num = request.POST.get('to_number')
            print(from_num, to_num)
            # for item in dir(request.POST):
            #     print(type(item), item)
            # print(dir(request.POST),type(request.POST.getlist))

    return redirect('/rubka')
